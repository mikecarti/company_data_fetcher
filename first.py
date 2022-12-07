#!/usr/bin/env python3
import os
import requests
import openpyxl
import pandas as pd
import yaml
import time
from dadata import Dadata

# prerequisites:
#
# python3.10
# openpyxl-3.0.10
# pandas-1.5.1
# pyyaml-6.0
# requests-2.28.1 urllib3-1.26.12
# dadata-21.10.1


# BASE_DIR = os.path.normpath(os.getcwd())
BASE_DIR = 'C:\\Users\\user\\Documents\\Python_Scripts\\virtual_envs\\venv_checko\\checko_project\\checko2'
EACH_COMPANY_LIMIT = 2
SAVE_EVERY_N_ITERATIONS = 25
FIRST_UNFILTERED_NUM = 3
FILTER_BY_MOSCOW = 3
FILTER_BY_CITY = 3
SECOND_SHEET = 'Все Данные с API'


def main() -> None:
    print('basedir: ', BASE_DIR)
    api_keys, company_names, file_name, url, SEARCH_COL_NAME = init()

    index = 0
    xlsx_write_file = os.path.join(BASE_DIR, file_name)
    workbook = openpyxl.load_workbook(xlsx_write_file)

    while True:
        if index > len(company_names) - 1:
            print(f'{index} row: OK')
            break
        name = get_nth_key(company_names, index)
        city = company_names[name]

        attempts = 0
        successful = False
        while attempts <= 10 and attempts < len(api_keys):
            key = api_keys[0 + attempts]

            data, err = request_to_api(url, key, name, index)
            attempts += 1

            if not err:
                write_data(workbook, path=xlsx_write_file, data=data, index=index, key=key, name=name, city=city)
                index += 1
                successful = True
                if attempts > 1:
                    for i in range(attempts - 1):
                        print(f"Ключ '{api_keys[0]}' истратил сегодняшние лимиты. Ключ больше не будет использоваться")
                        api_keys.pop(0)
                break
            else:
                print(f'{index + 1} (Attempt #{attempts}) with key: "{key}" - Fail')

        if not successful:
            print(f'{index + 1}: Ignore company with the name: "{name}". Data from api is corrupted!')
            index += 1

        if len(api_keys) == 0:
            print(f'No valid keys!')
            break
        if index % SAVE_EVERY_N_ITERATIONS == 0:
            workbook.save(xlsx_write_file)

        time.sleep(0.03)
    workbook.save(xlsx_write_file)


def collect_unused_data_to_xlsx(xlsx_write_file: str):
    global SEARCH_COL_NAME
    worksheets = pd.read_excel(xlsx_write_file, sheet_name=[0, 1])
    row_sheet = worksheets[0]
    clean_sheet = worksheets[1]

    # Д. Тимченко 03.11
    unused_sheet = pd.merge(row_sheet, clean_sheet.drop_duplicates(), on=SEARCH_COL_NAME, how='left'
                            , indicator=True).drop_duplicates(subset=[SEARCH_COL_NAME])
    unused_sheet = unused_sheet[unused_sheet['_merge'] == 'left_only']
    unused_sheet_only_names = unused_sheet[[SEARCH_COL_NAME]]

    workbook = openpyxl.load_workbook(xlsx_write_file)
    writer = pd.ExcelWriter(xlsx_write_file,
                            engine='openpyxl',
                            if_sheet_exists='overlay',
                            mode="a")
    writer.book = workbook
    print(unused_sheet_only_names)
    unused_sheet_only_names.to_excel(writer, sheet_name='Самые подходящие')
    writer.close()
    print("OK: Unused data is saved in sheet")


def get_city_from_adress(adress: str) -> str:
    # Возвращает символы после "г." и до "," или ".", без пробелов в начале и конце
    import re
    result = re.search('г (.*?)(\.|,)', adress)
    if result is None:
        return ""
    else:
        return result.group(1).strip()


def do_cities_match(adress: str, city_name: str):
    return get_city_from_adress(adress) == city_name


def write_clean_data_to_xlsx(workbook, data: list[dict], name: str, city_name: str, path: str) -> None:
    worksheet = workbook["Самые подходящие"]
    table_height = get_table_height(path, worksheet.title)
    # fill_worksheet(data, name, table_height, worksheet, city_name_filter=city_name)


def write_data(workbook, path, data: dict, index: int, key: str, name: str, city: str):
    print(f'{index + 1}: successful request to api with key: "{key}, name: "{name}"!')
    print(f'{index + 1}: write data to xlsx file with name: "{name}"')
    write_data_to_xlsx(workbook, data, name, city)

    # print(f'{index + 1}: clean data with name: "{name}"')
    # write_clean_data_to_xlsx(workbook, data, name, city, path=path)


def init():
    """
    Loads config and input data
    :return: api_keys: list, company_data: dict, file_name: str, url: str
    """
    config = yaml.load(
        stream=open(
            file=os.path.join(BASE_DIR, 'config.yml'),
            mode='r',
            encoding='utf-8'
        ),
        Loader=yaml.Loader
    )
    file_name = config['out_file']['xlsx_write_file']
    make_xlsx_out_file(
        base_dir=BASE_DIR,
        file_name=file_name,
        fields=config['out_file']['fields'].values()
    )
    company_data = load_company_data(
        os.path.join(BASE_DIR, config['xlsx_read_file'])
    )
    url = 'https://api.checko.ru/v2/search?key={key}&by=name&obj=org&query={name}&active=true'
    api_keys = config['api_keys']
    SEARCH_COL_NAME = config['col_names']['search_col_name']
    return api_keys, company_data, file_name, url, SEARCH_COL_NAME


def request_to_api(url: str, key: str, name: str, index: int) -> tuple[list[dict], bool]:
    print(f'{index + 1}: request to api with key: "{key}", name: "{name}"')

    response = requests.get(
        url=url.format(key=key, name=name.replace(' ', '+')),
    )
    print(response.reason)
    print(response.request)

    if response.status_code == 200:
        data = response.json()
        return data['data']['Записи'], False
    else:
        return {}, True


def request_to_api_dadata(url: str, key: str, name: str, index: int) -> tuple[list[dict], bool]:
    print(f'{index + 1}: request to api with key: "{key}", name: "{name}"')

    dadata = Dadata(key)

    result = dadata.suggest(name="party", query=name, count=1000, status=["ACTIVE"])

    f = lambda r: "No output from Api" if len(r) == 0 else 'Got output from Api'

    print(f(result))

    if len(result) == 0:
        return result, True
    else:
        return result, False


def write_data_to_xlsx(workbook, data: list[dict], name: str, city: str) -> None:
    worksheet = workbook[SECOND_SHEET]

    table_height = 1
    while True:
        if str(worksheet['A' + str(table_height)].value) in [None, 'None']:
            break
        else:
            table_height += 1

    load_rows_to_xlsx(city, data, name, table_height, worksheet)
    # load_rows_to_xlsx_dadata(city, data, name, table_height, worksheet)


def load_rows_to_xlsx(city, data, name, table_height, worksheet):
    for index, row in enumerate(data, start=table_height):
        worksheet['A' + str(index)].value = 0
        worksheet['B' + str(index)].value = do_cities_match(row['ЮрАдрес'], city_name=city)
        worksheet['C' + str(index)].value = name
        worksheet['D' + str(index)].value = row['ОГРН']
        worksheet['E' + str(index)].value = row['ИНН']
        worksheet['F' + str(index)].value = row['КПП']
        worksheet['G' + str(index)].value = row['НаимСокр']
        worksheet['H' + str(index)].value = row['НаимПолн']
        worksheet['I' + str(index)].value = row['ДатаРег']
        worksheet['J' + str(index)].value = row['Статус']
        worksheet['K' + str(index)].value = row['РегионКод']
        worksheet['L' + str(index)].value = row['ЮрАдрес']
        worksheet['M' + str(index)].value = row['ОКВЭД']
        worksheet['M' + str(index)].value = '________'


def load_rows_to_xlsx_dadata(city, data, name, table_height, worksheet):
    skipped_rows = 0
    new_rows = 0
    city_filter_bound = FIRST_UNFILTERED_NUM + FILTER_BY_CITY

    print(f'{name}: длинна {len(data)}')
    for index, row in enumerate(data, start=table_height):
        cur_data = row['data']

        if cur_data is None:
            continue

        cur_index = index - skipped_rows
        if FIRST_UNFILTERED_NUM <= new_rows < FIRST_UNFILTERED_NUM + FILTER_BY_CITY:
            # Filter by city
            if 'address' in cur_data and cur_data['address'] is not None:
                address = cur_data['address']['value'].lower()
                right_city = city.lower() in address
                print(f"Фильтрую по адресу: {address}, совпадение с городом {city} - {right_city}")
                if not right_city:
                    skipped_rows += 1
                    continue
        elif city_filter_bound < new_rows:
            print(f"Достигнут верхний лимит city фильтра")
            break
        else:
            address = cur_data['address']['value'].lower()
            print(f"Не фильтрую {address}")

        new_rows = load_to_excel_row(name, worksheet, new_rows, cur_data, cur_index)

    #  get rid of unfiltered rows
    data = data[FIRST_UNFILTERED_NUM:]
    # collect 'москва' matches again for that cycle without first unfiltered rows
    next_writing_point = new_rows + table_height
    new_rows = 0
    for index, row in enumerate(data, start=next_writing_point):
        cur_data = row['data']

        if cur_data is None:
            continue

        if new_rows < FILTER_BY_MOSCOW:
            # Filter by 'москва'
            if 'address' in cur_data and cur_data['address'] is not None:  # было  if 'adress'
                address = cur_data['address']['value'].lower()
                right_city = 'москва' in address or 'москов' in address
                print(f"Фильтрую по адресу: {address}, совпадение с москва - {right_city}")
                if not right_city:
                    skipped_rows += 1
                    continue
        elif new_rows >= FILTER_BY_MOSCOW:
            print(f"Достигнут верхний лимит 'москва' фильтра")
            break

        new_rows = load_to_excel_row(name, worksheet, new_rows, cur_data, cur_index)


def load_to_excel_row(name, worksheet, added_rows, cur_data, cur_index) -> int:
    worksheet['A' + str(cur_index)].value = 0

    worksheet['C' + str(cur_index)].value = name
    worksheet['D' + str(cur_index)].value = cur_data['ogrn']
    worksheet['E' + str(cur_index)].value = cur_data['inn']

    if 'name' in cur_data and cur_data['name'] is not None:
        worksheet['G' + str(cur_index)].value = cur_data['name']['short_with_opf']
        worksheet['H' + str(cur_index)].value = cur_data['name']['full_with_opf']
        worksheet['N' + str(cur_index)].value = cur_data['name']['short']

    if 'address' in cur_data and cur_data['address'] is not None:
        worksheet['L' + str(cur_index)].value = cur_data['address']['unrestricted_value']

    if 'management' in cur_data and cur_data['management'] is not None:
        worksheet['P' + str(cur_index)].value = cur_data['management']['post']
        worksheet['O' + str(cur_index)].value = cur_data['management']['name']

    if 'state' in cur_data and cur_data['state'] is not None:
        worksheet['J' + str(cur_index)].value = cur_data['state']['status']
    worksheet['M' + str(cur_index)].value = cur_data['okved']
    added_rows += 1

    return added_rows


def translit_names(path):
    pass


def create_named_sheet(path, fields, sheet_name=""):
    # if starting_sheet:
    #     workbook = openpyxl.Workbook()
    #     worksheet = workbook.active
    #     worksheet.title = "Все Данные с API"
    # else:
    workbook = openpyxl.load_workbook(path)

    if sheet_name not in workbook.sheetnames:
        worksheet = workbook.create_sheet(sheet_name)

        letters = ('A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P')
        for field, letter in zip(fields, letters):
            worksheet[letter + '1'].value = field
        workbook.save(path)
        workbook.close()
        print(f"Create worksheet with name '{worksheet.title}': OK")
    else:
        print("Unable to create " + sheet_name + ". Already exists")


def make_xlsx_out_file(
        base_dir: str,
        file_name: str,
        fields: list[str]
) -> None:
    """ Create xlsx file to write response data.

    :arg:
        base_dir: dir from this module.
        file_name: xlsx file name.
        fields: header fields to fill in xlsx file.
    """
    file_path = os.path.join(base_dir, file_name)

    if os.path.exists(file_path):
        create_named_sheet(file_path, fields, sheet_name=SECOND_SHEET)
        # create_named_sheet(file_path, fields, sheet_name="Самые подходящие")
        # create_named_sheet(file_path, fields, sheet_name="Не использованные")

    else:
        print("Ошибка, файл " + file_path + " не существует.")


def load_company_data(xlsx_path: str) -> list[str]:
    """ Read company names from xlsx file.

    :arg:
        xlsx_path: path to xlsx file.
    :returns:
        company_data: dict{name: str -> city: str}
    """
    company_data = {}
    workbook = openpyxl.load_workbook(xlsx_path)
    worksheet = workbook.active
    worksheet.title = "Входные данные"
    workbook.save(xlsx_path)

    table_height = get_table_height(xlsx_path, worksheet.title)
    for i in range(2, table_height):
        # Бывает нужно здесь изменить букву если ошибка 0 row: OK
        if worksheet['H' + str(i)].value != 1:
            continue
        company_name = str(worksheet['A' + str(i)].value)
        company_city = str(worksheet['B' + str(i)].value)
        if company_name != '0':
            company_data[company_name] = company_city
    return company_data


def get_table_height(path, sheet_name):
    excel_table = pd.read_excel(path, sheet_name)
    return len(excel_table) + 1


def get_nth_key(dictionary, n=0):
    if n < 0:
        n += len(dictionary)
    for i, key in enumerate(dictionary.keys()):
        if i == n:
            return key
    raise IndexError("dictionary index out of range")


if __name__ == '__main__':
    try:
        main()
    except KeyboardInterrupt:
        pass

# %%
