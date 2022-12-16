#!/usr/bin/env python3
import json
import os
import time

import requests
import openpyxl
import yaml
import pandas as pd
import first

BASE_DIR = os.path.normpath(os.getcwd())
JSONS_PATH = os.path.join(BASE_DIR, 'jsons/')
BUFFER_PATH = 'buffer.csv'
SECOND_SHEET = "Все Данные с API"
BAD_JSONS = False
MAX_REQUEST_PER_KEY = 99
API_MAX_ATTEMPTS = 2


def main() -> None:
    user_input = 0.00001
    variants = ('0', '1', '2', '3')

    print("""
    -Выберите режим
    1. Получить и распарсить данные по файлу Excel с инн
    2. Получить json'ы по файлу Excel с инн
    3. Распарсить данные из json'ов
    0. Отменить
    """)

    while user_input not in variants:
        user_input = input()
        if user_input not in variants:
            print("Неверный ввод, введите цифру от 0 до 3")

    match user_input:
        case '1':
            series = fetch_jsons_from_api()
            parse_jsons(series)
        case '2':
            fetch_jsons_from_api()
        case '3':
            series_list = collect_text_files_into_series()
            parse_jsons(series_list)
        case '0':
            return
        case _:
            raise ValueError("Ввод пользователя получил неверный аргумент")


def collect_text_files_into_series():
    series_list = []
    for filename in os.listdir(JSONS_PATH):
        file_path = JSONS_PATH + filename
        if os.path.isfile(file_path):
            series = pd.read_csv(file_path, sep='-=-=_+_=-=-=-+_+_', header=None, squeeze=True)
            series_list.append(series)
    return series_list


def parse_jsons(series_list):
    parsed_info_df = pd.DataFrame()

    status = [0, 0]
    if type(series_list) != list:
        series_list = [series_list]

    for s, series in enumerate(series_list):
        for i in range(series.size):
            json_string = series[i]
            try:
                json_dict = get_json_dict(json_string)
                res = 'Success'
                status[0] += 1
            except Exception as e:
                print(f"ОШИБКА: {e}")
                char_from = str(e).find('(') + 6
                char_to = str(e).find(')')
                char_at = int(str(e)[char_from:char_to])
                radius = 100
                print(
                    f"ОШИБКА JSON ТУТ, {char_at}: {json_string[char_at - radius:char_at + radius]}_"
                    f"{json_string[char_at:char_at + 1]}_{json_string[char_at + 1:char_at + radius]}")

                res = 'Error'
                status[1] += 1
            print(s, i + 1, ':', res)

            if res == 'Success':
                parsed_info_df = parse_json_into_df(json_dict, parsed_info_df)
    print(f'\nSuccessful: {status[0]}\n Failed: {status[1]}')

    parsed_info_df.to_csv('test.csv')

    writer = pd.ExcelWriter('данные_юр_лица.xlsx')
    parsed_info_df.to_excel(writer, index=False)
    writer.save()
    print('DataFrame is written successfully to Excel File.')


def fetch_jsons_from_api():
    config = yaml.load(
        stream=open(
            file=os.path.join(BASE_DIR, 'config.yml'),
            mode='r',
            encoding='utf-8'
        ),
        Loader=yaml.Loader
    )
    # out.xlsx file
    xlsx_load_file = config['out_file']['xlsx_write_file']
    # load inn from 'xlsx_write_file' (неверное кол-во, на один меньше)
    companies_inn = load_inn(xlsx_load_file, only_matched_city=False)
    url = 'https://api.checko.ru/v2/company?key={key}&inn={inn}&active=true'
    workbook = openpyxl.load_workbook(xlsx_load_file)
    worksheet = workbook[SECOND_SHEET]
    api_keys = config['api_keys']
    table_height = 2
    data_list = []

    # Начало цикла
    while True:
        if table_height - 1 > len(companies_inn):
            print(f'{table_height} row: OK')
            series = save_series(data_list)
            return series

        inn = companies_inn[table_height - 2]
        successful = False

        # -1 because of header in excel file
        if not need_to_find_inn(table_height, worksheet):
            print(f"{table_height}: ИНН '{inn}' не помечен на дальнейший отбор. Проверяю далее")
            table_height += 1
            continue
        print(f"{table_height}: ИНН '{inn}' помечен на дальнейший отбор. Запрашиваю данные")

        successful, table_height = try_getting_data(api_keys, inn, successful, table_height, url, data_list)

        if not successful:
            print(f'{table_height}: Ignore company with the INN: "{inn}".'
                  f' Data from api is corrupted!')
            table_height += 1
            continue

        if table_height % 100 == 0:
            save_series(data_list)


def save_series(data_dict):
    s = pd.Series(data_dict)
    s_decoded = s.apply(lambda x: json.loads(x))
    s.to_csv(BUFFER_PATH + '_encoded', index=False)
    s_decoded.to_csv(BUFFER_PATH, index=False)
    return s


def try_getting_data(api_keys, inn, successful, table_height, url, data_list):
    attempts = 0
    while attempts <= len(api_keys) and attempts <= API_MAX_ATTEMPTS:
        key = api_keys[0 + attempts]

        print(f'{table_height}: request to api with key: "{key}", inn: "{inn}"')
        data, err = request_to_api(url, key, inn)
        attempts += 1

        if not err:
            successful = True
            table_height = write_data(api_keys, data, inn, key, table_height, data_list)
            break
        else:
            print(f'{table_height} (Attempt #{attempts}) with key: "{key}" - Fail')
    return successful, table_height


def need_to_find_inn(table_height: int, ws):
    return str(ws['A' + str(table_height)].value) in ("1", "True")


def write_data(api_keys, data, inn, key, table_height, data_list):
    print(f'{table_height}: successful request to api with key: "{key}, inn: "{inn}"!')
    cur_key_valid = append_json_to_list(data, data_list)
    table_height += 1
    validate_key(api_keys, cur_key_valid)
    return table_height


def validate_key(api_keys, cur_key_valid):
    if not cur_key_valid:
        print(f"Key {api_keys[0]} is fully used and deleted from stack.")
        api_keys.pop(0)


def request_to_api(url: str, key: str, inn: str) -> tuple[dict, bool]:
    response = requests.get(
        url=url.format(key=key, inn=inn),
    )
    if response.status_code == 200:
        data = response.json()
        return data, False
    else:
        return {}, True


def get_column_height(path, sheet_name):
    excel_table = pd.read_excel(path, sheet_name)
    company_columns = ['УпрОрг ИНН', 'УпрОрг НаимПолн', 'УстКап Сумма', 'Руковод', 'Контакты Телефон',
                       'Контакты email', 'Контакты ВебСайт', 'СЧР', 'this_key_requests']
    company_data = excel_table[company_columns]
    if company_data.last_valid_index():
        return company_data.last_valid_index() + 1
    else:
        return 2


def append_json_to_list(data: dict, data_list: list) -> bool:
    data_str = json.dumps(data)
    data_list.append(data_str)
    return data['meta']['today_request_count'] < MAX_REQUEST_PER_KEY


def load_inn(xlsx_path: str, only_matched_city=False) -> list[str]:
    companies_inn = []
    workbook = openpyxl.load_workbook(xlsx_path)
    worksheet = workbook[SECOND_SHEET]
    table_height = first.get_table_height(xlsx_path, worksheet.title)
    for i in range(2, table_height + 1):
        if only_matched_city:
            if str(worksheet['B' + str(i)].value).lower() in ("false", "0", "нет"):
                continue

        company_inn = str(worksheet['E' + str(i)].value)
        if company_inn not in ['0', None, 'None']:
            companies_inn.append(company_inn)
    return companies_inn


def get_json_dict(json_string):
    assert type(json_string) == str, f"type: {type(json_string)}"
    if BAD_JSONS:
        json_string = json_string.replace('\"', '').replace('None', '\'None\''). \
            replace('False', '\'ЛОЖЬ\'').replace('True', '\'ПРАВДА\'').replace('\'', '\"'). \
            replace('\xa0', ' ')
    return json.loads(json_string)


def get_field_data(nested_data, field):
    field_data_result = ''
    field_data = nested_data[field]
    if not field_data or field_data == 'None':
        return ''

    if type(field_data) == list:
        for info in nested_data[field]:
            field_data_result += f"{info} ; "
    else:
        field_data_result += f"{field_data} ; "
    return field_data_result


def get_inn(json_dict):
    if json_dict['ИНН']:
        return json_dict['ИНН']


def get_name(json_dict):
    # print(json_dict)
    if json_dict['Статус']:
        return json_dict['Статус']['Наим']


def get_adress(json_dict):
    # print(json_dict)
    if json_dict['ЮрАдрес']:
        return json_dict['ЮрАдрес']['АдресРФ']


def get_person_info(json_dict):
    # print(json_dict)
    if not json_dict['Руковод']:
        return ''

    heads = ""
    for person in json_dict['Руковод']:
        heads += f"{person['ФИО']} ; {person['ИНН']} ; {person['НаимДолжн']}\n"
    return heads


def collect_data(data_dict, on_fields: list):
    if not data_dict:
        return ''

    data = ''
    for field in on_fields:
        data += get_field_data(data_dict, field)
    return data


def get_uchred_fl(json_dict):
    if not json_dict['Учред'] or not json_dict['Учред']['ФЛ']:
        return

    individual_list = json_dict['Учред']['ФЛ']
    data = ''
    for person in individual_list:
        data += person['ФИО'] + ' ; '
        data += person['ИНН'] + ' ; '
        if person['Доля']:
            data += process_percent(person)
        data += ' \n'
    return data


def get_russian_individuals(json_dict):
    if not json_dict['Учред'] and not json_dict['Учред']['РосОрг']:
        return

    russian_individuals = json_dict['Учред']['РосОрг']
    data = ''
    for person in russian_individuals:
        data += person['ИНН'] + ' ; '
        data += person['НаимПолн'] + ' ; '
        if person['Доля']:
            data += process_percent(person)
        data += ' \n'
    return data


def process_percent(person):
    percent = person['Доля']['Процент']
    if percent == 'None' or not percent:
        return 'None'
    else:
        percent_f = float(percent)
        return str(round(percent_f, 3))


def get_connected_individuals(json_dict):
    connected_individuals = json_dict['СвязУпрОрг']
    data = ''
    for person in connected_individuals:
        data += f"{person['НаимСокр']} ; {person['ИНН']}\n"
    return data


def get_subsidiary_num(json_dict):
    if not json_dict['Подразд']:
        return 0

    subsidiaries = json_dict['Подразд']['Филиал']
    return len(subsidiaries) if len(subsidiaries) > 0 else "None"


def get_representatives_num(json_dict):
    if not json_dict['Подразд']:
        return 0

    representatives = json_dict['Подразд']['Представ']
    return len(representatives) if len(representatives) > 0 else "None"


def parse_json_into_df(json_dict, df):
    # 1. +++Статус:Наим+++
    # 2. +++ЮрАдрес:АдресРФ+++
    # 3. +++УпрОрг:ИНН; УпрОрг:НаимПолн+++
    # 4. +++Руковод:ФИО;Руковод:ИНН;Руковод:НаимДолжности+++
    # 5. +++[Учред:ФЛ:ФИО ; Учред:ФЛ:ИНН ; Учред:ФЛ:Доля:Процент]+++
    # 6. +++[Учред:РосОрг:ИНН ; Учред:РосОрг:НаимПолн и Учред:РосОрг:Доля:Процент]+++
    # 7. +++[СвязУпрОрг:НаимСокр; СвязУпрОрг:ИНН]+++
    # 8. +++Подразд:Филиал - количество записей+++
    # 9. +++Подразд:Представ - количество записей+++
    # 10. +++СЧР+++
    # 11. +++Контакты:Тел; Контакты:Емэйл; Контакты:ВебСайт+++

    json_dict = json_dict['data']
    # print(f"ЛОГ:  {json_dict}")
    if json_dict:
        row = {
            'ИНН': get_inn(json_dict),
            'Статус': get_name(json_dict),
            'АдресРФ': get_adress(json_dict),
            'УпрОрг': collect_data(json_dict['УпрОрг'], on_fields=['ИНН', 'НаимПолн']),
            'Руковод (ФИО_ИНН_НаимДолжности)': get_person_info(json_dict),
            'Учред ФЛ (ФИО_ИНН_Процент)': get_uchred_fl(json_dict),
            'Учред (ИНН_НаимПолн_Процент)': get_russian_individuals(json_dict),
            'СвязУпрОрг': get_connected_individuals(json_dict),
            'Кол-во Филиалов': get_subsidiary_num(json_dict),
            'Кол-во Представителей': get_representatives_num(json_dict),
            'СЧР': json_dict['СЧР'],
            'Тел_Емэйл_Вебсайт': collect_data(json_dict['Контакты'], on_fields=['Тел', 'Емэйл', 'ВебСайт'])
        }
        print(f"ДАННЫЕ: {row}")

        df = pd.concat([df, pd.DataFrame(row, index=[0])], ignore_index=True)
    return df


if __name__ == '__main__':
    try:
        main()
    except KeyboardInterrupt:
        pass
